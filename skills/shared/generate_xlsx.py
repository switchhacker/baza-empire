#!/usr/bin/env python3
"""
Baza Empire Skill — generate_xlsx
Generate an Excel (.xlsx) spreadsheet from data.
Phil uses this for invoices, estimates, budgets, financial summaries, schedules, trackers.

SKILL_ARGS:
  title       (str)   — Spreadsheet title / first sheet name
  sheets      (list)  — [{name: str, headers: [...], rows: [[...], ...]}, ...]
  filename    (str)   — output filename (default: title_snake.xlsx)
  project_id  (str)   — dashboard project (default: shared)
  summary_row (bool)  — add SUM row at bottom of numeric columns (default: false)
"""
import os
import sys
import json
import re
import time
import urllib.request

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: venv/bin/pip install openpyxl"}))
    sys.exit(1)

DASHBOARD_URL = os.environ.get("BAZA_DASHBOARD_URL", "http://localhost:8888")
ARTIFACTS_BASE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "dashboard", "artifacts"
)

args = json.loads(os.environ.get("SKILL_ARGS", "{}"))

title      = args.get("title", "Spreadsheet")
sheets     = args.get("sheets", [])
project_id = args.get("project_id", "shared")
add_sum    = args.get("summary_row", False)

raw_name = args.get("filename", "")
if not raw_name:
    safe = re.sub(r'[^\w]', '_', title.lower())[:40]
    raw_name = f"{safe}_{int(time.time())}.xlsx"
if not raw_name.endswith(".xlsx"):
    raw_name += ".xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
EVEN_FILL    = PatternFill("solid", fgColor="EBF3FF")
SUM_FONT     = Font(bold=True, size=11)
thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default empty sheet

# If no sheets provided, create one default sheet from top-level keys
if not sheets:
    # Support flat format: title + headers + rows at top level
    sheets = [{
        "name": title[:30],
        "headers": args.get("headers", []),
        "rows": args.get("rows", []),
    }]

for sheet_def in sheets:
    sheet_name = sheet_def.get("name", "Sheet")[:31]
    headers    = sheet_def.get("headers", [])
    rows       = sheet_def.get("rows", [])

    ws = wb.create_sheet(title=sheet_name)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=sheet_def.get("title", sheet_name))
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Header row (row 2)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = border
    ws.row_dimensions[2].height = 18

    # Data rows
    for row_idx, row_data in enumerate(rows, 3):
        fill = EVEN_FILL if (row_idx % 2 == 0) else None
        for col_idx, val in enumerate(row_data[:len(headers)], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if fill:
                cell.fill = fill
            # Auto-format numbers starting with $
            if isinstance(val, str) and val.startswith("$"):
                try:
                    cell.value = float(val.replace("$","").replace(",",""))
                    cell.number_format = '"$"#,##0.00'
                except:
                    pass

    # Summary (SUM) row
    if add_sum and rows:
        sum_row = len(rows) + 3
        for col_idx, header in enumerate(headers, 1):
            # Try to detect numeric columns from first data row
            first_val = rows[0][col_idx - 1] if rows and (col_idx - 1) < len(rows[0]) else None
            is_numeric = isinstance(first_val, (int, float)) or (
                isinstance(first_val, str) and first_val.startswith("$")
            )
            if is_numeric:
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=sum_row, column=col_idx,
                               value=f"=SUM({col_letter}3:{col_letter}{sum_row-1})")
                cell.font = SUM_FONT
                cell.number_format = '"$"#,##0.00'
                cell.border = border
            else:
                cell = ws.cell(row=sum_row, column=col_idx,
                               value="TOTAL" if col_idx == 1 else "")
                cell.font = SUM_FONT

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_data in rows:
            if (col_idx - 1) < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

out_dir = os.path.join(ARTIFACTS_BASE, project_id)
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, raw_name)
wb.save(out_path)

# Register with dashboard
try:
    boundary = "bazaxlsxboundary"
    with open(out_path, "rb") as f:
        file_data = f.read()
    body = (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"project_id\"\r\n\r\n{project_id}\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"{raw_name}\"\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_data + f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(
        f"{DASHBOARD_URL}/api/artifacts/upload",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=15):
        pass
except Exception:
    pass

print(json.dumps({
    "success": True,
    "filename": raw_name,
    "path": out_path,
    "project_id": project_id,
    "sheets": len(sheets),
    "download_url": f"{DASHBOARD_URL}/api/artifacts/download/{project_id}/{raw_name}",
}))
